from __future__ import annotations

import math
from collections import defaultdict
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from scipy.stats import rankdata, friedmanchisquare, wilcoxon
from sklearn.metrics import cohen_kappa_score

APPROACH_ORDER = ["Prompting", "RAFG", "Finetuning", "DRAFT"]
APPROACH_POSITIONS = ["Model_A", "Model_B", "Model_C", "Model_D"]
METRICS = ["closeness", "correctness"]


def load_evaluator_rows(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    """Load evaluation rows from sheets containing a task identifier column."""
    sheet_rows: dict[str, list[dict[str, Any]]] = {}

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                continue

            headers = [
                str(cell).strip() if cell is not None else ""
                for cell in rows[0]
            ]

            identifier_header = next(
                (
                    header
                    for header in ("Decision_ID", "Task_ID")
                    if header in headers
                ),
                None,
            )

            if identifier_header is None:
                continue

            parsed_rows = []

            for row in rows[1:]:
                if not has_content(row):
                    continue

                parsed_row = {
                    header: row[index] if index < len(row) else None
                    for index, header in enumerate(headers)
                }
                parsed_row["Decision_ID"] = safe_float(parsed_row[identifier_header])
                parsed_rows.append(parsed_row)

            sheet_rows[worksheet.title] = parsed_rows

    finally:
        workbook.close()

    return sheet_rows


def has_content(row: tuple[Any, ...]) -> bool:
    """Return True when a row contains at least one non-empty value."""
    return any(
        cell is not None and str(cell).strip()
        for cell in row
    )


def safe_float(value: Any) -> float:
    """Convert a value to float, treating missing/empty values as NaN."""
    if value is None or str(value).strip() == "":
        return np.nan
    try:
        return float(value)
    except ValueError:
        return np.nan


def get_approach_score(
    row: dict[str, Any],
    position: str,
) -> tuple[str, float, float] | None:
    """Return (approach, closeness, correctness) for an approach position."""
    approach = row.get(position)

    if approach is None or not str(approach).strip():
        return None

    approach = str(approach).strip()

    closeness = safe_float(row.get(f"{position}_Closeness"))
    correctness = safe_float(row.get(f"{position}_Correctness"))

    return approach, closeness, correctness


def get_approach_scores(row: dict[str, Any]) -> list[tuple[str, float, float]]:
    """Return scores for all populated approach positions in a row."""
    scores = []

    for position in APPROACH_POSITIONS:
        score = get_approach_score(row, position)
        if score is not None:
            scores.append(score)

    return scores


def build_reliability_matrices(
    evaluators: dict[str, list[dict[str, Any]]]
) -> tuple[np.ndarray, np.ndarray]:
    """
    Aligns scores by Decision_ID and Approach, converting raw scores into 
    rankings for each sample to calculate agreement on relative preferences.
    """
    evaluator_keys = list(evaluators.keys())
    num_evaluators = len(evaluator_keys)

    closeness_data = defaultdict(lambda: [np.nan] * num_evaluators)
    correctness_data = defaultdict(lambda: [np.nan] * num_evaluators)

    for i, eval_key in enumerate(evaluator_keys):
        for row in evaluators[eval_key]:
            decision_id = row.get("Decision_ID")
            if not decision_id or math.isnan(decision_id):
                continue

            scores = get_approach_scores(row)
            if not scores:
                continue

            approaches = [s[0] for s in scores]
            closeness_vals = np.array([s[1] for s in scores])
            correctness_vals = np.array([s[2] for s in scores])

            def get_ranks(vals: np.ndarray) -> np.ndarray:
                ranks = np.full(len(vals), np.nan)
                valid_mask = ~np.isnan(vals)
                if np.any(valid_mask):
                    valid_vals = vals[valid_mask]
                    valid_ranks = rankdata([-v for v in valid_vals], method='average')
                    ranks[valid_mask] = valid_ranks
                return ranks

            closeness_ranks = get_ranks(closeness_vals)
            correctness_ranks = get_ranks(correctness_vals)

            for approach, c_rank, corr_rank in zip(approaches, closeness_ranks, correctness_ranks):
                key = (decision_id, approach)
                closeness_data[key][i] = c_rank
                correctness_data[key][i] = corr_rank

    closeness_matrix = np.array(list(closeness_data.values())).T
    correctness_matrix = np.array(list(correctness_data.values())).T

    return closeness_matrix, correctness_matrix


def build_raw_matrices_for_kappa(
    evaluators: dict[str, list[dict[str, Any]]]
) -> tuple[np.ndarray, np.ndarray, list[str]]:
    """
    Aligns raw Likert scores (unranked) for pairwise Cohen's Kappa evaluation.
    """
    evaluator_keys = list(evaluators.keys())
    num_evaluators = len(evaluator_keys)

    closeness_data = defaultdict(lambda: [np.nan] * num_evaluators)
    correctness_data = defaultdict(lambda: [np.nan] * num_evaluators)

    for i, eval_key in enumerate(evaluator_keys):
        for row in evaluators[eval_key]:
            decision_id = row.get("Decision_ID")
            if not decision_id or math.isnan(decision_id):
                continue

            scores = get_approach_scores(row)
            for approach, closeness, correctness in scores:
                key = (decision_id, approach)
                closeness_data[key][i] = closeness
                correctness_data[key][i] = correctness

    closeness_matrix = np.array(list(closeness_data.values())).T
    correctness_matrix = np.array(list(correctness_data.values())).T

    return closeness_matrix, correctness_matrix, evaluator_keys



def calculate_pairwise_kappas(matrix: np.ndarray, evaluator_keys: list[str]) -> list[str]:
    """Calculate Weighted Cohen's Kappa for all rater pairs."""
    results = []
    num_raters = matrix.shape[0]
    
    for i in range(num_raters):
        for j in range(i + 1, num_raters):
            r1_name = evaluator_keys[i].split(" (")[0]
            r2_name = evaluator_keys[j].split(" (")[0]
            
            # Keep only samples where BOTH raters provided a score
            valid_mask = ~np.isnan(matrix[i]) & ~np.isnan(matrix[j])
            r1 = matrix[i][valid_mask]
            r2 = matrix[j][valid_mask]
            
            if len(r1) == 0:
                results.append(f"  * {r1_name} vs {r2_name}: N/A (no overlap)")
                continue
                
            # Convert to integers (Cohen's Kappa expects discrete categories)
            r1 = np.round(r1).astype(int)
            r2 = np.round(r2).astype(int)
            
            try:
                # Quadratic weights heavily penalize large mismatches (e.g., 1 vs 5)
                kappa = cohen_kappa_score(r1, r2, weights="quadratic")
                results.append(f"  * {r1_name} vs {r2_name}: {kappa:.3f} (n={len(r1)})")
            except Exception as e:
                results.append(f"  * {r1_name} vs {r2_name}: Error ({e})")
                
    return results


def summarize_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Calculate approach means and rankings, ignoring NaN values."""
    approach_stats = defaultdict(
        lambda: {
            "closeness": [],
            "correctness": [],
        }
    )

    for row in rows:
        scores = get_approach_scores(row)
        if not scores:
            continue

        for approach, closeness, correctness in scores:
            if not math.isnan(closeness):
                approach_stats[approach]["closeness"].append(closeness)
            if not math.isnan(correctness):
                approach_stats[approach]["correctness"].append(correctness)

    mean_per_approach = {}

    for approach in APPROACH_ORDER:
        if approach not in approach_stats:
            continue

        c_vals = approach_stats[approach]["closeness"]
        corr_vals = approach_stats[approach]["correctness"]

        mean_per_approach[approach] = {
            "closeness": sum(c_vals) / len(c_vals) if c_vals else np.nan,
            "correctness": sum(corr_vals) / len(corr_vals) if corr_vals else np.nan,
        }

    return {
        "mean_per_approach": {
            approach: {
                "closeness": round(stats["closeness"], 3) if not math.isnan(stats["closeness"]) else "N/A",
                "correctness": round(stats["correctness"], 3) if not math.isnan(stats["correctness"]) else "N/A",
            }
            for approach, stats in mean_per_approach.items()
        }
    }


def combined_mean(
    rows_list: list[dict[str, Any]],
) -> dict[str, dict[str, float | str]]:
    """Calculate approach means across multiple evaluator datasets."""
    scores = defaultdict(
        lambda: {
            "closeness": [],
            "correctness": [],
        }
    )

    for row in rows_list:
        for approach, closeness, correctness in get_approach_scores(row):
            if not math.isnan(closeness):
                scores[approach]["closeness"].append(closeness)
            if not math.isnan(correctness):
                scores[approach]["correctness"].append(correctness)

    combined = {}

    for approach in APPROACH_ORDER:
        if approach not in scores:
            continue

        closeness_values = scores[approach]["closeness"]
        correctness_values = scores[approach]["correctness"]

        combined[approach] = {
            "closeness": round(sum(closeness_values) / len(closeness_values), 3) if closeness_values else "N/A",
            "correctness": round(sum(correctness_values) / len(correctness_values), 3) if correctness_values else "N/A",
        }

    return combined


def get_averaged_author_scores(evaluators: dict[str, list[dict[str, Any]]], metric: str) -> dict[str, list[float]]:
    """Averages scores for the two authors across identical samples for statistical testing."""
    author_keys = [k for k in evaluators.keys() if "Evaluator" in k]
    if len(author_keys) != 2:
        return {}

    decision_scores = defaultdict(lambda: defaultdict(list))

    for author in author_keys:
        for row in evaluators[author]:
            dec_id = row.get("Decision_ID")
            if not dec_id or math.isnan(dec_id):
                continue
            for approach, closeness, correctness in get_approach_scores(row):
                val = closeness if metric == "closeness" else correctness
                if not math.isnan(val):
                    decision_scores[dec_id][approach].append(val)

    approach_lists = {app: [] for app in APPROACH_ORDER}

    for dec_id, app_dict in decision_scores.items():
        if all(app in app_dict and len(app_dict[app]) > 0 for app in APPROACH_ORDER):
            for app in APPROACH_ORDER:
                avg_score = sum(app_dict[app]) / len(app_dict[app])
                approach_lists[app].append(avg_score)

    return approach_lists


def run_statistical_tests(evaluators: dict[str, list[dict[str, Any]]], target: str = "DRAFT") -> list[str]:
    """Runs Friedman omnibus test followed by Wilcoxon signed-rank with Holm-Bonferroni."""
    lines = []

    for metric in METRICS:
        lines.append(f"\n## Statistical Significance: {metric.capitalize()}")
        data = get_averaged_author_scores(evaluators, metric)

        if not data or not all(len(v) > 0 for v in data.values()):
            lines.append("*Not enough perfectly paired data to run statistical tests across all 4 approaches.*")
            continue

        n_samples = len(data[APPROACH_ORDER[0]])
        arrays = [data[app] for app in APPROACH_ORDER]
        
        stat, p_friedman = friedmanchisquare(*arrays)
        
        lines.append(f"* **Matched Sample Size:** {n_samples}")
        lines.append(f"* **Friedman Test (Omnibus):** p = {p_friedman:.4f}")

        if p_friedman >= 0.05:
            lines.append("* *Conclusion:* No statistically significant difference found among the 4 approaches (p >= 0.05). Pairwise tests omitted.")
            continue

        lines.append(f"* *Conclusion:* Significant difference detected. Proceeding to pairwise tests against {target}.")
        lines.append(f"\n### Pairwise Wilcoxon Tests vs. {target} (Holm-Bonferroni Corrected)")

        target_data = data[target]
        baselines = [app for app in APPROACH_ORDER if app != target]
        raw_p_values = []
        
        for baseline in baselines:
            baseline_data = data[baseline]
            diffs = [t - b for t, b in zip(target_data, baseline_data)]
            if all(d == 0 for d in diffs):
                raw_p_values.append(1.0)
            else:
                _, p_wilc = wilcoxon(target_data, baseline_data)
                raw_p_values.append(p_wilc)

        sorted_indices = np.argsort(raw_p_values)
        m = len(raw_p_values)
        adj_p_values = np.zeros(m)

        for i, idx in enumerate(sorted_indices):
            adj_p_values[idx] = raw_p_values[idx] * (m - i)
            if i > 0:
                adj_p_values[idx] = max(adj_p_values[idx], adj_p_values[sorted_indices[i-1]])
            adj_p_values[idx] = min(adj_p_values[idx], 1.0)

        for idx in range(m):
            baseline = baselines[idx]
            adj_p = adj_p_values[idx]
            sig_marker = "**Significant**" if adj_p < 0.05 else "Not Significant"
            
            mean_target = sum(target_data) / n_samples
            mean_baseline = sum(data[baseline]) / n_samples
            
            winner = target if mean_target > mean_baseline else (baseline if mean_baseline > mean_target else "Tie")
            lines.append(f"  * **{target} vs. {baseline}:** p = {adj_p:.4f} ({sig_marker}) | *Higher Mean: {winner}*")

    return lines


def find_evaluator_sheets(
    base_dir: Path,
    task: str
) -> dict[str, list[dict[str, Any]]]:
    """Load the two author evaluators."""
    authors = load_evaluator_rows(base_dir / f"{task} Authors.xlsx")

    author_rows = list(authors.values())

    if len(author_rows) < 2:
        raise ValueError(
            "Need at least two evaluator sheets in the Authors workbook.")

    return {
        f"Evaluator 1 ({list(authors.keys())[0]})": author_rows[0],
        f"Evaluator 2 ({list(authors.keys())[1]})": author_rows[1],
    }


def format_results(
    evaluators: dict[str, list[dict[str, Any]]],
) -> str:
    """Generate the complete evaluation report."""
    summaries = {
        name: summarize_rows(rows)["mean_per_approach"]
        for name, rows in evaluators.items()
    }

    all_rows = [row for rows in evaluators.values() for row in rows]
    summaries["Combined"] = combined_mean(all_rows)

    lines = [
        "## Performance Means",
        "| Evaluator | Approach | Closeness | Correctness |",
        "|---|---|---:|---:|"
    ]

    for evaluator, summary in summaries.items():
        for approach in APPROACH_ORDER:
            if approach in summary:
                lines.append(
                    f"| {evaluator} | {approach} | "
                    f"{summary[approach]['closeness']} | "
                    f"{summary[approach]['correctness']} |"
                )
    # Adding pairwise Weighted Cohen's Kappa
    lines.append("\n## Pairwise Inter-Rater Agreement (Weighted Cohen's Kappa)")
    lines.append("*(Uses raw Likert scores. Quadratic weights applied to penalize severe disagreements.)*")
    
    raw_close, raw_corr, eval_keys = build_raw_matrices_for_kappa(evaluators)
    
    lines.append("\n**Closeness Pairwise:**")
    lines.extend(calculate_pairwise_kappas(raw_close, eval_keys))
    
    lines.append("\n**Correctness Pairwise:**")
    lines.extend(calculate_pairwise_kappas(raw_corr, eval_keys))

    lines.append("\n*(Note for Kappa: <0=Poor, .01-.20=Slight, .21-.40=Fair, .41-.60=Moderate, .61-.80=Substantial, .81-1=Almost Perfect)*")

    lines.extend(run_statistical_tests(evaluators, target="DRAFT"))

    return "\n".join(lines)


def visualize_raw_matrix(
    evaluators: dict[str, list[dict[str, Any]]], 
    metric: str = "closeness"
) -> str:
    """Creates a formatted text table of raw scores for all evaluators."""
    evaluator_keys = list(evaluators.keys())
    
    matrix_data = defaultdict(lambda: ["NaN"] * len(evaluator_keys))
    
    for i, eval_key in enumerate(evaluator_keys):
        for row in evaluators[eval_key]:
            decision_id = row.get("Decision_ID")
            if not decision_id or math.isnan(decision_id):
                continue
            
            for approach, closeness, correctness in get_approach_scores(row):
                key = (str(decision_id), approach)
                val = closeness if metric == "closeness" else correctness
                
                if math.isnan(val):
                    matrix_data[key][i] = "NaN"
                else:
                    matrix_data[key][i] = f"{val:.1f}"

    def sort_key(k):
        dec_id, app = k
        app_idx = APPROACH_ORDER.index(app) if app in APPROACH_ORDER else 999
        return (float(dec_id) if dec_id.replace('.','',1).isdigit() else dec_id, app_idx)
        
    sorted_keys = sorted(matrix_data.keys(), key=sort_key)
    
    headers = ["Decision_ID", "Approach"] + evaluator_keys
    col_widths = [len(h) for h in headers]
    for k in sorted_keys:
        col_widths[0] = max(col_widths[0], len(k[0]))
        col_widths[1] = max(col_widths[1], len(k[1]))
    col_widths = [w + 2 for w in col_widths]
    
    lines = []
    lines.append(f"### Raw Scores Matrix: {metric.capitalize()}")
    
    header_line = "".join(h.ljust(w) for h, w in zip(headers, col_widths))
    lines.append(header_line)
    lines.append("-" * len(header_line))
    
    for k in sorted_keys:
        row_data = [k[0], k[1]] + matrix_data[k]
        row_line = "".join(str(item).ljust(w) for item, w in zip(row_data, col_widths))
        lines.append(row_line)
        
    return "\n".join(lines)


def main(task: str) -> None:
    base_dir = Path(__file__).resolve().parent
    output_file = base_dir / f"{task}_results.txt"

    try:
        evaluators = find_evaluator_sheets(base_dir, task)
        
        # print(f"\n{'='*60}")
        # print(f"VISUALIZING DATA FOR TASK: {task}")
        # print(f"{'='*60}")
        # print(visualize_raw_matrix(evaluators, metric="closeness"))
        # print("\n")
        # print(visualize_raw_matrix(evaluators, metric="correctness"))
        # print(f"{'='*60}\n")
        
        report = format_results(evaluators)

        output_file.write_text(
            report,
            encoding="utf-8",
        )
        print(f"[{task}] Results written to: {output_file}")

    except Exception as e:
        print(f"[{task}] Failed to process: {e}")


if __name__ == "__main__":
    main("CD")
    main("TB")